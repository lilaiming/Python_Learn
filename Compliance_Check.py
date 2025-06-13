# -*- coding=utf-8 -*-
# 本脚由lilaiming编写，用于学习使用！
# Email:essid@qq.com

import os
import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def parse_source_text(source_text):
    sections = {}
    current_section = None
    lines = source_text.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 检测是否是新的章节标题 (数字开头)
        if re.match(r'^\d+\.\d+\.\d+', line):
            current_section = line
            sections[current_section] = []
        elif current_section:
            sections[current_section].append(line)

    return sections


def check_target_compliance(source_sections, target_content):
    results = {}
    # 将目标文本转换为行列表（保留原始内容）
    target_lines = [line.rstrip('\n') for line in target_content.splitlines()]

    for section, source_lines in source_sections.items():
        compliant = True

        # 逐行检查源文本中的每一行
        for source_line in source_lines:
            found = False
            for target_line in target_lines:
                # 严格比较每一行内容
                if source_line.strip() == target_line.strip():
                    found = True
                    break

            # 如果任何一行缺失或不匹配
            if not found:
                compliant = False
                break

        results[section] = "Compliance" if compliant else "Non-compliance"

    return results


def main():
    # 定义源文本内容
    source_text = """
3.1.2 TIMESTAMP
info-center timestamp debugging format-date precision-time tenth-second
info-center timestamp log format-date precision-time tenth-second
3.1.3 PASSWORD ENCRYPTION
paremerg@local
parfunct@local
3.1.5 DHCP SERVICE
undo dhcp enable
3.1.7 TIME ZONE
clock timezone HKT add 08:00:00
3.1.8 VTP MODE
vcmp role transparent
stp mode vbst
3.1.10 GRATUITOUS ADDRESS RESOLUTION PROTOCOL (ARP)
undo arp gratuitous-arp send enable
3.1.11 DOMAIN NAME SERVICE
undo dns resolve
dns domain hkairport.com
3.1.12 SPANNING TREE FEATURES
stp bpdu-protection
stp edged-port default
3.1.13 MANAGEMENT INTERFACE
interface Vlanif2
3.1.14 HTTP and HTTPS SERVERS
undo http server enable
undo http secure-server enable
3.1.15 LOGGING
info-center source default channel 0 log state off trap state off
info-center source default channel 1 log level informational
info-center source default channel 2 log level notification trap state off
info-center source default channel 4 log level informational
info-center loghost 168.106.16.6 facility local6 local-time
info-center loghost 168.106.16.7 facility local6 local-time
info-center logbuffer size 1024
3.1.16 SIMPLE NETWORK MANAGEMENT PTOTOCOL (SNMP)
snmp-agent trap disable
snmp-agent trap enable feature-name ERROR-DOWN
snmp-agent trap enable feature-name VCMP
snmp-agent trap enable feature-name LACP
snmp-agent trap enable feature-name DLDP
snmp-agent trap enable feature-name ENTITYMIB
snmp-agent sys-info version v2c v3
snmp-agent usm-user v3 anms-nac
3.1.17 BANNER
header login information
header shell information
3.1.18 REMOTE ACCESS CONTROL
stelnet server enable
ssh client first-time enable
user-interface con 0
authentication-mode aaa
user-interface vty 0 4
authentication-mode aaa
idle-timeout 5 0
3.1.19 NETWORK TIME PROTOCAL (NTP)
ntp-service unicast-server 168.106.194.254 preference
ntp-service unicast-server 168.106.194.253
3.1.21 Authentication, Authorization and Accounting
hwtacacs-server template aanet
hwtacacs-server authentication 168.106.18.175
hwtacacs-server authentication 168.106.18.176 secondary
3.2.1 INTERFACE DESCRIPTION
No description
3.2.2 SWITCH PORT MODE
port link-type
3.2.3 SPANNING TREE FEATURES
stp loop-protection
3.2.4 MAXIMUM TRANSMISSION UNIT (MTU)
mtu 9198
3.2.5 UNIDIRECTIONAL LINK DETECTION (UDLD)
dldp enable
3.2.6 LAYER 3 INTERFACE MESSAGES
undo icmp redirect send
undo icmp host-unreachable send
undo arp-proxy enable
3.2.7 DISABLE UNUSED INTERFACE
shutdown
3.2.8 PORT SECURITY
port-security enable
port-security protect-action shutdown
port-security mac-address sticky
3.2.9 BROADCAST STORM CONTROL
storm-control broadcast min-rate percent 2 max-rate percent 2
storm-control multicast min-rate percent 2 max-rate percent 2
3.2.10 MULTICAST CONFIGURATION
igmp-snooping enable
igmp-snooping fast-switch enable
igmp version 3
igmp ssm-mapping enable
igmp-snooping prompt-leave
statistic enable
3.2.11 QOS configuration
qos pq 4 to 7 drr 0 to 3
3.3.2 ACL naming
Non-compliance
    """.strip()

    # 解析源文本
    source_sections = parse_source_text(source_text)
    section_headers = list(source_sections.keys())

    # 创建Tkinter根窗口并隐藏
    root = tk.Tk()
    root.withdraw()

    # 选择目标文件夹
    target_folder = filedialog.askdirectory(title="选择目标文件夹")

    if not target_folder:
        print("未选择文件夹，程序退出")
        return

    # 获取文件夹中的所有文本文件（支持.txt和.log）
    target_files = []
    for root_dir, _, files in os.walk(target_folder):
        for file in files:
            if file.lower().endswith(('.txt', '.log')):
                target_files.append(os.path.join(root_dir, file))

    if not target_files:
        print("文件夹中没有找到文本文件(.txt/.log)")
        return

    print(f"找到 {len(target_files)} 个文本文件")

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "合规性报告"

    # 设置标题行
    headers = ["目标文件"] + section_headers
    ws.append(headers)

    # 设置样式
    compliance_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    non_compliance_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 处理每个目标文件
    for file_path in target_files:
        try:
            # 读取文件内容
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                target_content = f.read()

            # 检查合规性
            results = check_target_compliance(source_sections, target_content)
            filename = os.path.basename(file_path)

            # 准备行数据
            row = [filename] + [results[header] for header in section_headers]
            ws.append(row)

            # 设置单元格样式
            row_idx = ws.max_row
            for col_idx, value in enumerate(row[1:], start=2):  # 从第二列开始
                cell = ws.cell(row=row_idx, column=col_idx)
                if value == "Compliance":
                    cell.fill = compliance_fill
                else:
                    cell.fill = non_compliance_fill

            print(f"已处理: {filename}")
        except Exception as e:
            print(f"处理文件 {os.path.basename(file_path)} 时出错: {e}")

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 保存Excel文件
    output_file = os.path.join(target_folder, "合规性报告.xlsx")
    wb.save(output_file)
    print(f"结果已保存到: {output_file}")
    print("操作完成")


if __name__ == "__main__":
    main()
